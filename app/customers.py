# app/customers.py
# قراءة وعرض بيانات العملاء من Excel

from openpyxl import load_workbook

# مسار ملف العملاء
EXCEL_FILE = "data/customers.xlsx"

# فتح الملف
wb = load_workbook(EXCEL_FILE)
ws = wb.active

# عنوان
print("=" * 70)
print("📋 قائمة عملاء شركة المصاعد")
print("=" * 70)

# عداد العملاء
count = 0

# قراءة الصفوف
for row in ws.iter_rows(min_row=2, values_only=True):
    count += 1

    print(f"\n🏢 عميل #{count}")
    print(f"📌 كود العميل:      {row[0]}")
    print(f"👤 اسم العميل:      {row[2]}")
    print(f"📄 رقم العقد:       {row[3]}")
    print(f"📍 المدينة:         {row[4]}")
    print(f"📍 الحي:            {row[5]}")
    print(f"📱 الجوال:          {row[7]}")
    print(f"🛗 عدد المصاعد:     {row[9]}")
    print(f"⚡ الحالة:          {row[11]}")
    print(f"💰 الإيراد:         {row[14]} ريال")

    # عرض أول 5 فقط
    if count >= 5:
        break

# ملخص
print("\n" + "=" * 70)
print(f"📊 تم عرض أول {count} عملاء")
print("=" * 70)