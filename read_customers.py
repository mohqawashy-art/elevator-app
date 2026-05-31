# read_customers.py
# سكريبت لقراءة ملف العملاء وعرضهم

from openpyxl import load_workbook

# 1. نفتح ملف العملاء (الموجود)
wb = load_workbook("customers.xlsx")

# 2. ناخد الشيت الأول
ws = wb.active

# 3. نطبع رأس مزخرف
print("=" * 60)
print("📋 قائمة العملاء")
print("=" * 60)

# 4. نقرا الصفوف واحد واحد ونعرضهم
# min_row=2 معناها ابدأ من الصف رقم 2 (نتخطى صف العناوين)
# values_only=True معناها جيب القيم بس (مش معلومات الخلية)
for row in ws.iter_rows(min_row=2, values_only=True):
    customer_id, name, phone, location = row
    print(f"🔹 {customer_id} | {name} | الجوال: {phone} | الموقع: {location}")

print("=" * 60)
print(f"✅ إجمالي العملاء: {ws.max_row - 1}")