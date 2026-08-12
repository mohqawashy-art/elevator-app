#!/usr/bin/env python3
"""استيراد عملاء من Excel إلى مستأجر (مثل jama على PostgreSQL).

  cd ~/liftcore/elevator-app
  set -a; source /etc/liftcore/platform.env; set +a
  python scripts/import_clients_xlsx_tenant.py /path/to/clients.xlsx --slug jama --dry-run
  python scripts/import_clients_xlsx_tenant.py /path/to/clients.xlsx --slug jama --yes
"""
from __future__ import annotations

import argparse
import os
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit('pip install openpyxl') from exc


def load_rows(path: str) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    sheet_name = 'العملاء' if 'العملاء' in wb.sheetnames else wb.sheetnames[-1]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    out: list[dict] = []
    for row in rows[2:]:  # تخطّي صف التلميحات
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue
        item = {}
        for h, v in zip(headers, row):
            if not h:
                continue
            item[h] = '' if v is None else str(v).strip()
        if not (item.get('الاسم (عربي)') or item.get('اسم العميل') or item.get('الاسم')):
            continue
        # تجاهل صف التلميحات إن وُضع في البيانات
        name = item.get('الاسم (عربي)') or ''
        if name.startswith('إلزامي') or 'يُولَّد' in (item.get('رقم العميل') or ''):
            continue
        out.append(item)
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description='استيراد عملاء Excel لمستأجر')
    parser.add_argument('xlsx', help='مسار ملف Excel')
    parser.add_argument('--slug', default='jama')
    parser.add_argument('--dry-run', action='store_true')
    parser.add_argument('--yes', action='store_true', help='تنفيذ الاستيراد فعلياً')
    args = parser.parse_args()

    if not args.dry_run and not args.yes:
        print('ERROR: حدّد --dry-run أو --yes')
        return 2

    path = os.path.abspath(args.xlsx)
    if not os.path.isfile(path):
        print(f'ERROR: الملف غير موجود: {path}')
        return 1

    rows = load_rows(path)
    print(f'==> ملف: {path}')
    print(f'==> صفوف صالحة: {len(rows)}')
    if not rows:
        print('ERROR: لا توجد صفوف للاستيراد')
        return 1

    from app import app
    from client_bulk_import import import_customer_rows
    from flask import g
    from models import Customer, Organization

    slug = (args.slug or 'jama').strip().lower()

    with app.app_context():
        org = Organization.query.filter_by(slug=slug).first()
        if not org:
            print(f'ERROR: لا توجد مؤسسة slug={slug}')
            return 1
        g.organization = org
        g.organization_id = org.id
        before = (
            Customer.query.execution_options(skip_tenant=True)
            .filter_by(organization_id=org.id)
            .count()
        )
        print(f'==> مستأجر: {org.slug} (id={org.id}) — عملاء حاليون: {before}')

        if args.dry_run:
            from client_bulk_import import normalize_import_row, _prepare_import_phone
            from app import client_phone_error
            from form_validation import customer_name_error

            ok = 0
            bad = 0
            for i, raw in enumerate(rows, start=1):
                data = normalize_import_row(raw)
                name = data.get('name') or ''
                phone = _prepare_import_phone(data.get('phone') or '')
                err = customer_name_error(name) or client_phone_error(phone)
                if err:
                    bad += 1
                    if bad <= 10:
                        print(f'  صف {i}: {err} | {name!r} | {phone!r}')
                else:
                    ok += 1
            print(f'DRY-RUN: صالح={ok} مشاكل={bad} (بدون حفظ)')
            return 0 if bad == 0 else 1

        result = import_customer_rows(rows)
        after = (
            Customer.query.execution_options(skip_tenant=True)
            .filter_by(organization_id=org.id)
            .count()
        )
        print(f"==> imported={result.get('imported')} failed={result.get('failed')}")
        print(f'==> عملاء بعد: {after} (كان {before})')
        for err in (result.get('errors') or [])[:15]:
            print(f"  خطأ صف {err.get('row')}: {err.get('error')}")
        return 0 if result.get('failed', 0) == 0 else 1


if __name__ == '__main__':
    raise SystemExit(main())
