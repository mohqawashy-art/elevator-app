#!/usr/bin/env python3
"""إنشاء قالب استيراد العملاء على سطح المكتب."""
from __future__ import annotations

import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def desktop_path() -> str:
    if sys.platform.startswith('win'):
        home = os.environ.get('USERPROFILE') or os.path.expanduser('~')
        return os.path.join(home, 'Desktop')
    return os.path.join(os.path.expanduser('~'), 'Desktop')


def build(out_path: str) -> str:
    wb = Workbook()

    # ورقة البيانات أولاً حتى يقرأ الاستيراد في الواجهة الورقة الصحيحة
    ws = wb.active
    ws.title = 'العملاء'
    ws.sheet_view.rightToLeft = True

    headers = [
        ('رقم العميل', 'اختياري — مثل C-0001؛ اتركه فارغاً ليُولَّد تلقائياً'),
        ('الاسم (عربي)', 'إلزامي — اسم العميل كما سيظهر في النظام'),
        ('رقم الهاتف', 'إلزامي — 05xxxxxxxx أو 5xxxxxxxx أو +9665xxxxxxxx'),
        ('المدينة', 'مثل: مكة المكرمة / جدة'),
        ('الحي', 'الحي أو المنطقة'),
        ('العنوان', 'العنوان التفصيلي للموقع / الخريطة'),
        ('البريد الإلكتروني', 'اختياري'),
        ('اسم المسؤول', 'شخص التواصل'),
        ('نوع المتعاقد', 'فرد أو شركة فقط'),
        ('رقم هوية المتعاقد', 'للأفراد — رقم الهوية الوطنية'),
        ('رقم السجل التجاري', 'للشركات فقط'),
        ('الرقم الضريبي', 'للشركات — مطلوب للفواتير الضريبية'),
        ('العنوان الوطني', 'للشركات — عنوان وطني للفواتير'),
        ('حالة العميل', 'نشط أو غير نشط'),
        ('ملاحظات', 'اختياري'),
    ]

    header_fill = PatternFill('solid', fgColor='0F3D68')
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    hint_fill = PatternFill('solid', fgColor='EEF4FA')
    hint_font = Font(name='Calibri', color='4A5568', size=9, italic=True)
    thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )

    for col, (h, hint) in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin
        hcell = ws.cell(2, col, hint)
        hcell.fill = hint_fill
        hcell.font = hint_font
        hcell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        hcell.border = thin

    for r in range(3, 53):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c, '')
            cell.border = thin

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 42
    widths = [14, 28, 16, 16, 14, 32, 24, 16, 14, 18, 18, 18, 18, 12, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    dv_entity = DataValidation(type='list', formula1='"فرد,شركة"', allow_blank=True)
    ws.add_data_validation(dv_entity)
    dv_entity.add('I3:I200')

    dv_status = DataValidation(type='list', formula1='"نشط,غير نشط"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add('N3:N200')

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    ins = wb.create_sheet('تعليمات')
    ins.sheet_view.rightToLeft = True
    ins['A1'] = 'قالب استيراد عملاء جما — LiftCore'
    ins['A1'].font = Font(name='Calibri', bold=True, size=16, color='0F3D68')
    lines = [
        '',
        'عبّئ ورقة «العملاء» فقط — صف واحد لكل عميل.',
        'الصف 1 عناوين (لا تغيّرها). الصف 2 تلميحات. ابدأ البيانات من الصف 3.',
        '',
        'إلزامي: الاسم (عربي) + رقم الهاتف',
        'الجوال يقبل: 05xxxxxxxx أو 5xxxxxxxx أو +9665xxxxxxxx',
        'فرد → رقم هوية | شركة → سجل تجاري + رقم ضريبي + عنوان وطني',
    ]
    for i, line in enumerate(lines, 2):
        ins[f'A{i}'] = line
        ins[f'A{i}'].font = Font(name='Calibri', size=12)
    ins.column_dimensions['A'].width = 100

    os.makedirs(os.path.dirname(out_path) or '.', exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == '__main__':
    target = os.path.join(desktop_path(), 'قالب-استيراد-عملاء-جما.xlsx')
    if len(sys.argv) > 1:
        target = sys.argv[1]
    path = build(target)
    print(f'[OK] {path} ({os.path.getsize(path)} bytes)')
