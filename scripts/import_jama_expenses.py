#!/usr/bin/env python3
"""Import Jama expenses from Excel with salary split and type normalization.

  python scripts/import_jama_expenses.py deploy/data/jama_import/expenses_11_7_2026.xlsx --slug jama --dry-run
  python scripts/import_jama_expenses.py deploy/data/jama_import/expenses_11_7_2026.xlsx --slug jama
"""
from __future__ import annotations

import argparse
import os
import re
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from openpyxl import load_workbook

from app import app, db
from import_real_data import _cell, _f, _i, _parse_date, _str
from models import Expense
from tenant_scope import assign_organization, tenant_query

# توزيع الرواتب الشهرية (يُستخدم كنسب من إجمالي دفعة الرواتب في Excel)
SALARY_LINES: list[tuple[str, float]] = [
    ('رواتب موظفي الأعطال', 7000.0),
    ('راتب فني الصيانة', 2500.0),
    ('رواتب المساعد', 1300.0),
    ('راتب الإداري', 2500.0),
    ('راتب المدير', 3500.0),
]
SALARY_WEIGHT_SUM = sum(w for _, w in SALARY_LINES)

TYPE_MAP = {
    'محروقات': 'محروقات',
    'وقود': 'محروقات',
    'شراء قطع غيار': 'قطع غيار',
    'قطع غيار': 'قطع غيار',
    'صيانه سيارات': 'صيانة سيارات',
    'صيانة سيارات': 'صيانة سيارات',
    'ادوات مكتبية': 'أدوات',
    'أدوات': 'أدوات',
    'اخرى': 'أخرى',
    'أخرى': 'أخرى',
    'ضيافة': 'ضيافة',
    'مصاريف صيانة': 'مصاريف صيانة',
    'مصاريف تجديد وتحديث مصاعد': 'مصاريف تجديد وتحديث مصاعد',
    'نقل': 'نقل',
    'مصروفات اساسية': 'مصروفات أساسية',
    'مصروفات أساسية': 'مصروفات أساسية',
}


def _is_lump_salary(notes: str, raw_type: str) -> bool:
    text = f'{notes} {raw_type}'
    return bool(re.search(r'رواتب|راتب', text))


def _split_salary_total(total: float) -> list[tuple[str, float]]:
    """قسّم إجمالي الرواتب الشهري حسب أوزان جما."""
    total = round(_f(total), 2)
    if total <= 0:
        return []
    parts: list[tuple[str, float]] = []
    allocated = 0.0
    for idx, (label, weight) in enumerate(SALARY_LINES):
        if idx == len(SALARY_LINES) - 1:
            amt = round(total - allocated, 2)
        else:
            amt = round(total * weight / SALARY_WEIGHT_SUM, 2)
            allocated += amt
        parts.append((label, amt))
    return parts


def _normalize_expense_type(raw_type: str, notes: str) -> str:
    raw = _str(raw_type)
    notes_l = _str(notes).lower()
    if _is_lump_salary(notes, raw):
        return 'رواتب'
    mapped = TYPE_MAP.get(raw, raw or 'أخرى')
    if mapped in ('مصروفات أساسية', 'مصروفات اساسية'):
        if 'راتب' in notes_l or 'رواتب' in notes_l:
            return 'رواتب'
        return 'مصروفات أساسية'
    return mapped or 'أخرى'


def _salary_codes(base_num: int) -> list[str]:
    return [f'EXP-{base_num:04d}-{i:02d}' for i in range(1, len(SALARY_LINES) + 1)]


def _purge_salary_group(base_num: int, *, dry_run: bool) -> None:
    codes = [f'EXP-{base_num:04d}'] + _salary_codes(base_num)
    for code in codes:
        row = tenant_query(Expense).filter_by(code=code).first()
        if row and not dry_run:
            db.session.delete(row)


def _build_expense(
    *,
    code: str,
    edate,
    expense_type: str,
    description: str,
    responsible: str,
    payment_method: str,
    amount: float,
    reference: str,
    notes: str,
) -> Expense:
    expense = Expense(
        code=code,
        expense_date=edate,
        expense_type=expense_type,
        description=description,
        responsible=responsible,
        payment_method=payment_method or 'كاش',
        amount=round(_f(amount), 2),
        reference=reference,
        notes=notes,
    )
    assign_organization(expense)
    return expense


def _load_rows(path: str) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return []
    keys = [(_str(h) or f'col_{i}') for i, h in enumerate(header)]
    out: list[dict] = []
    for row in rows_iter:
        if not any(v is not None and _str(v) for v in row):
            continue
        item = {keys[i]: (row[i] if i < len(row) else None) for i in range(len(keys))}
        out.append(item)
    return out


def import_expenses(
    path: str,
    *,
    dry_run: bool = False,
    skip_existing: bool = True,
    sync_existing: bool = False,
) -> dict:
    rows = _load_rows(path)
    stats = {
        'rows': len(rows),
        'imported': 0,
        'updated': 0,
        'salary_splits': 0,
        'skipped_existing': 0,
        'errors': 0,
        'excel_total': 0.0,
        'types': {},
    }
    existing_codes = {e.code.upper() for e in tenant_query(Expense).all() if e.code}

    for r in rows:
        num = _i(_cell(r, 'رقم العملية'))
        base_code = f'EXP-{num:04d}' if num else ''
        edate = _parse_date(_cell(r, 'التاريخ'))
        amount = _f(_cell(r, 'المبلغ'))
        notes = _str(_cell(r, 'ملاحظات'))
        raw_type = _str(_cell(r, 'نوع المصروف'))
        responsible = _str(_cell(r, 'مسئول الصرف'))
        payment_method = _str(_cell(r, 'طريقة الدفع')) or 'كاش'
        reference = _str(_cell(r, 'مرفقات'))[:500]
        vendor = _str(_cell(r, 'المورد'))[:100]

        if not base_code or not edate or amount <= 0:
            stats['errors'] += 1
            continue

        stats['excel_total'] = round(stats['excel_total'] + amount, 2)

        is_salary_lump = _is_lump_salary(notes, raw_type)
        target_codes = _salary_codes(num) if is_salary_lump else [base_code]

        if sync_existing and is_salary_lump:
            _purge_salary_group(num, dry_run=dry_run)

        if is_salary_lump:
            if skip_existing and not sync_existing and all(
                c.upper() in existing_codes for c in target_codes
            ):
                stats['skipped_existing'] += len(target_codes)
                continue
        else:
            one = tenant_query(Expense).filter_by(code=base_code).first()
            if one and sync_existing:
                pass
            elif skip_existing and base_code.upper() in existing_codes:
                stats['skipped_existing'] += 1
                continue

        if is_salary_lump:
            splits = _split_salary_total(amount)
            for idx, (label, part_amount) in enumerate(splits):
                code = target_codes[idx]
                exp_type = 'رواتب'
                desc = label
                stats['types'][exp_type] = stats['types'].get(exp_type, 0) + 1
                existing_row = tenant_query(Expense).filter_by(code=code).first()
                fields = dict(
                    expense_date=edate,
                    expense_type=exp_type,
                    description=desc,
                    responsible=responsible,
                    payment_method=payment_method,
                    amount=part_amount,
                    reference=reference,
                    notes=vendor or notes,
                )
                if existing_row and sync_existing:
                    for k, v in fields.items():
                        setattr(existing_row, k, v)
                    action = 'updated'
                else:
                    expense = _build_expense(
                        code=code,
                        edate=edate,
                        expense_type=exp_type,
                        description=desc,
                        responsible=responsible,
                        payment_method=payment_method,
                        amount=part_amount,
                        reference=reference,
                        notes=vendor or notes,
                    )
                    if not dry_run:
                        db.session.add(expense)
                    action = 'imported'
                stats[action] += 1
                stats['salary_splits'] += 1
                if not dry_run:
                    existing_codes.add(code.upper())
            continue

        exp_type = _normalize_expense_type(raw_type, notes)
        desc = (notes or raw_type or exp_type)[:300]
        stats['types'][exp_type] = stats['types'].get(exp_type, 0) + 1

        fields = dict(
            expense_date=edate,
            expense_type=exp_type,
            description=desc,
            responsible=responsible,
            payment_method=payment_method,
            amount=amount,
            reference=reference,
            notes=vendor,
        )
        existing_row = tenant_query(Expense).filter_by(code=base_code).first()
        if existing_row and sync_existing:
            for k, v in fields.items():
                setattr(existing_row, k, v)
            stats['updated'] += 1
        else:
            expense = _build_expense(
                code=base_code,
                edate=edate,
                expense_type=exp_type,
                description=desc,
                responsible=responsible,
                payment_method=payment_method,
                amount=amount,
                reference=reference,
                notes=vendor,
            )
            if not dry_run:
                db.session.add(expense)
            stats['imported'] += 1
        if not dry_run:
            existing_codes.add(base_code.upper())

    if not dry_run:
        db.session.commit()

    stats['db_total'] = (
        round(sum(_f(e.amount) for e in tenant_query(Expense).all()), 2) if not dry_run else None
    )
    stats['db_salary_total'] = (
        round(
            sum(_f(e.amount) for e in tenant_query(Expense).filter_by(expense_type='رواتب').all()),
            2,
        )
        if not dry_run
        else None
    )
    return stats


def main() -> int:
    parser = argparse.ArgumentParser(description='Import Jama expenses from Excel')
    parser.add_argument('xlsx', help='Path to expenses .xlsx')
    parser.add_argument('--slug', default='jama', help='Organization slug')
    parser.add_argument('--dry-run', action='store_true')
    parser.add_argument('--force', action='store_true', help='Import even if expense code exists')
    parser.add_argument('--sync', action='store_true', help='Update existing rows and refresh salary splits')
    args = parser.parse_args()

    if not os.path.isfile(args.xlsx):
        print(f'ERROR: file not found: {args.xlsx}')
        return 1

    from flask import g
    from models import Organization

    with app.app_context():
        slug = (args.slug or 'jama').strip().lower()
        org = Organization.query.filter_by(slug=slug).first()
        if not org:
            print(f'ERROR: لا توجد مؤسسة slug={slug!r}')
            return 1
        g.organization = org
        g.organization_id = org.id
        print(f'Tenant: {org.name} ({org.slug})')
        print('File:', args.xlsx)
        result = import_expenses(
            args.xlsx,
            dry_run=args.dry_run,
            skip_existing=not (args.force or args.sync),
            sync_existing=args.sync,
        )
        print(result)
        if not args.dry_run:
            print('expenses in tenant:', tenant_query(Expense).count())
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
