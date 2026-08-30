#!/usr/bin/env python3
"""Import Jama revenues from Excel with contract / parts / invoice linking.

  python scripts/import_jama_revenues.py deploy/data/jama_import/revenues_11_7_2026.xlsx --slug jama --dry-run
  python scripts/import_jama_revenues.py deploy/data/jama_import/revenues_11_7_2026.xlsx --slug jama
"""
from __future__ import annotations

import argparse
import os
import re
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from openpyxl import load_workbook

from app import app, db, sync_contract_invoice_status
from customer_billing import (
    apply_payment_to_source,
    parts_remaining,
    repair_contract_payment_links,
    resolve_contract_id,
    split_vat_amounts,
)
from entity_links import contract_by_code, customer_by_name
from import_real_data import _cell, _extract_cn, _f, _i, _parse_date, _str
from models import Contract, Customer, Invoice, PartsBilling, Revenue
from tenant_scope import assign_organization, tenant_query

PARTS_REVENUE_TYPES = frozenset({'قطع غيار', 'بيع قطع غيار', 'زيارة'})
CONTRACT_REVENUE_TYPES = frozenset({'تجديد عقد', 'عقد صيانة', 'عقد جديد', 'ضمان'})

# ملف إيرادات جما: عمود «المبلغ» = الإجمالي المحصّل (شامل الضريبة)
JAMA_EXCEL_AMOUNT_IS_INCLUSIVE = True


def _normalize_cn(code: str | None) -> str | None:
    raw = _extract_cn(code or '') or _extract_cn(_str(code))
    if not raw:
        return None
    m = re.search(r'CN-(\d+)', raw.upper())
    if not m:
        return raw.upper()
    return f'CN-{int(m.group(1)):05d}'


def _customer_name_from_title(title: str) -> str:
    text = _str(title)
    text = re.sub(r'CN-\s*\d+\s*', '', text, flags=re.I).strip(' -|،,')
    return text


def _map_status(raw: str) -> str:
    s = _str(raw)
    if s in ('محصل', 'محصّل', 'مكتملة'):
        return 'محصّل'
    if s in ('معلق', 'معلّق'):
        return 'معلق'
    if 'لغ' in s:
        return 'ملغي'
    return s or 'محصّل'


def _normalize_revenue_type(raw: str) -> str:
    s = _str(raw)
    if not s:
        return 'أخرى'
    if 'بيع' in s and 'قطع' in s:
        return 'قطع غيار'
    if s == 'زيارة':
        return 'أعمال إضافية'
    if s in ('تجديد عقد', 'عقد جديد', 'قطع غيار', 'عقد صيانة', 'أعمال إضافية', 'أخرى'):
        return s
    if 'تجديد' in s or 'صيانة' in s:
        return 'تجديد عقد'
    if 'عقد' in s and 'جديد' in s:
        return 'عقد جديد'
    if 'قطع' in s:
        return 'قطع غيار'
    return s


def _amounts_from_excel(raw_amount: float) -> tuple[float, float, float]:
    """تحويل مبلغ Excel إلى (قبل الضريبة، الضريبة، الإجمالي)."""
    val = _f(raw_amount)
    if val <= 0:
        return 0.0, 0.0, 0.0
    if JAMA_EXCEL_AMOUNT_IS_INCLUSIVE:
        return split_vat_amounts(total_incl_vat=val)
    return split_vat_amounts(amount_ex_vat=val)


def _normalize_ar_name(name: str) -> str:
    s = ' '.join((name or '').strip().split())
    for src, dst in (('أ', 'ا'), ('إ', 'ا'), ('آ', 'ا'), ('ى', 'ي'), ('ة', 'ه')):
        s = s.replace(src, dst)
    return s


def _find_customer_by_title(title: str) -> Customer | None:
    name = _customer_name_from_title(title)
    if not name:
        return None
    exact = customer_by_name(name)
    if exact:
        return exact
    compact = _normalize_ar_name(name)
    for cust in tenant_query(Customer).all():
        cn = _normalize_ar_name(cust.name or '')
        if not cn:
            continue
        if cn == compact or compact in cn or cn in compact:
            return cust
    return tenant_query(Customer).filter(Customer.name.ilike(f'%{name[:20]}%')).first()


def _money_close(a: float, b: float, tol: float = 1.0) -> bool:
    return abs(_f(a) - _f(b)) <= tol


def _find_parts_billing(
    *,
    customer_id: int,
    contract_id: int | None,
    revenue_date,
    amount_ex: float,
    total_incl: float,
    used_parts: set[int],
) -> PartsBilling | None:
    q = tenant_query(PartsBilling).filter_by(customer_id=customer_id)
    if contract_id:
        q = q.filter(
            (PartsBilling.contract_id == contract_id) | (PartsBilling.contract_id.is_(None))
        )
    candidates = q.order_by(PartsBilling.billing_date.desc()).all()
    scored: list[tuple[int, PartsBilling]] = []

    for pb in candidates:
        if pb.id in used_parts:
            continue
        if pb.billing_date and revenue_date:
            days = abs((pb.billing_date - revenue_date).days)
            if days > 90:
                continue
        else:
            days = 999

        sell = _f(pb.sell_price)
        sell_incl = round(sell * 1.15, 2)
        remaining = parts_remaining(pb)

        score = 0
        if _money_close(amount_ex, sell) or _money_close(total_incl, sell) or _money_close(total_incl, sell_incl):
            score += 50
        elif _money_close(remaining, total_incl) or _money_close(remaining, amount_ex):
            score += 40
        elif remaining > 0.01:
            score += 10

        if contract_id and pb.contract_id == contract_id:
            score += 20
        if days <= 7:
            score += 15
        elif days <= 30:
            score += 8

        if score >= 40:
            scored.append((score, pb))

    if not scored:
        return None
    scored.sort(key=lambda x: (-x[0], x[1].billing_date or revenue_date), reverse=False)
    return scored[0][1]


def _find_invoice(
    *,
    customer_id: int,
    contract_id: int | None,
    total_incl: float,
    used_invoices: set[int],
) -> Invoice | None:
    q = tenant_query(Invoice).filter_by(customer_id=customer_id)
    if contract_id:
        q = q.filter(
            (Invoice.contract_id == contract_id) | (Invoice.contract_id.is_(None))
        )
    for inv in q.order_by(Invoice.invoice_date.desc()).all():
        if inv.id in used_invoices:
            continue
        if _money_close(inv.total, total_incl) or _money_close(inv.amount, total_incl / 1.15):
            return inv
    return None


def _resolve_customer_contract(row: dict) -> tuple:
    title = _str(_cell(row, 'Title'))
    contracts_col = _str(_cell(row, 'العقود'))
    cn = _normalize_cn(contracts_col) or _normalize_cn(title)
    contract = contract_by_code(cn) if cn else None
    customer = contract.customer if contract else None

    if not customer:
        for src in (contracts_col, title):
            if not src:
                continue
            name = _customer_name_from_title(src)
            if not name:
                continue
            customer = customer_by_name(name) or _find_customer_by_title(src)
            if customer:
                break

    # لا نربط بعقد آخر للعميل إذا كان رقم عقد محدداً في الملف ولم يُوجد
    if not contract and customer and not cn:
        contract = (
            tenant_query(Contract).filter_by(customer_id=customer.id)
            .order_by(Contract.end_date.desc())
            .first()
        )

    return contract, customer, cn


def _append_note(notes: str, extra: str) -> str:
    extra = (extra or '').strip()
    if not extra:
        return notes or ''
    if extra in (notes or ''):
        return notes or extra
    return f'{extra} — {notes}'.strip(' —') if notes else extra


def _load_rows(path: str) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return []
    keys = [(_str(h) or f'col_{i}') for i, h in enumerate(header)]
    out: list[dict] = []
    for row in rows_iter:
        if not any(v is not None and _str(v) for v in row):
            continue
        item = {keys[i]: (row[i] if i < len(row) else None) for i in range(len(keys))}
        out.append(item)
    return out


def import_revenues(
    path: str,
    *,
    dry_run: bool = False,
    skip_existing: bool = True,
    sync_existing: bool = False,
    import_orphans: bool = False,
) -> dict:
    rows = _load_rows(path)
    stats = {
        'rows': len(rows),
        'imported': 0,
        'updated': 0,
        'skipped_existing': 0,
        'skipped_missing': 0,
        'imported_no_contract': 0,
        'imported_orphan': 0,
        'errors': 0,
        'linked_contract': 0,
        'linked_parts': 0,
        'linked_invoice': 0,
        'excel_total': 0.0,
        'types': {},
    }
    missing_samples: list[str] = []
    contract_ids: set[int] = set()
    used_parts: set[int] = set()
    used_invoices: set[int] = set()

    existing_codes = {r.code.upper() for r in tenant_query(Revenue).all() if r.code}

    for r in rows:
        num = _i(_cell(r, 'رقم العملية'))
        code = f'REV-{num:04d}' if num else ''
        rdate = _parse_date(_cell(r, 'التاريخ'))
        raw_amount = _f(_cell(r, 'المبلغ'))
        amount_ex, tax, total_incl = _amounts_from_excel(raw_amount)
        stats['excel_total'] = round(stats['excel_total'] + raw_amount, 2)
        rev_type = _normalize_revenue_type(_cell(r, 'نوع الايراد'))

        if not code or not rdate or total_incl <= 0:
            stats['errors'] += 1
            continue

        existing_rev = tenant_query(Revenue).filter_by(code=code).first()
        if existing_rev and sync_existing:
            pass
        elif skip_existing and code.upper() in existing_codes:
            stats['skipped_existing'] += 1
            continue

        contract, customer, cn = _resolve_customer_contract(r)
        title = _str(_cell(r, 'Title'))
        contracts_col = _str(_cell(r, 'العقود'))
        if not contract and not customer:
            if not import_orphans:
                stats['skipped_missing'] += 1
                if len(missing_samples) < 25:
                    missing_samples.append(f'{code}: لا عقد/عميل لـ {cn or title or contracts_col}')
                continue
            stats['imported_orphan'] += 1

        customer_id = customer.id if customer else (contract.customer_id if contract else None)
        contract_id = contract.id if contract else None
        note_bits = []
        if cn and not contract_id:
            note_bits.append(f'عقد Excel: {cn}')
        if import_orphans and not customer_id:
            label = title or contracts_col
            if label:
                note_bits.append(f'عميل Excel: {label}')
        note_extra = ' | '.join(note_bits)
        row_notes = _append_note(_str(_cell(r, 'ملاحظات')), note_extra)

        if not contract_id and customer_id:
            contract_id = resolve_contract_id(
                customer_id,
                _str(_cell(r, 'مرفقات')),
                _str(_cell(r, 'ملاحظات')),
                cn or '',
                rev_type,
            )

        invoice_id = None
        parts_billing_id = None

        if rev_type in PARTS_REVENUE_TYPES:
            pb = _find_parts_billing(
                customer_id=customer_id,
                contract_id=contract_id,
                revenue_date=rdate,
                amount_ex=amount_ex,
                total_incl=total_incl,
                used_parts=used_parts,
            )
            if pb:
                parts_billing_id = pb.id
                if not contract_id and pb.contract_id:
                    contract_id = pb.contract_id
                stats['linked_parts'] += 1
                if not dry_run:
                    used_parts.add(pb.id)

        elif rev_type == 'عقد جديد':
            inv = _find_invoice(
                customer_id=customer_id,
                contract_id=contract_id,
                total_incl=total_incl,
                used_invoices=used_invoices,
            )
            if inv:
                invoice_id = inv.id
                if not contract_id and inv.contract_id:
                    contract_id = inv.contract_id
                stats['linked_invoice'] += 1
                if not dry_run:
                    used_invoices.add(inv.id)

        stats['types'][rev_type] = stats['types'].get(rev_type, 0) + 1

        if contract_id:
            stats['linked_contract'] += 1
            contract_ids.add(contract_id)
        elif customer_id and cn:
            stats['imported_no_contract'] += 1

        fields = dict(
            customer_id=customer_id,
            contract_id=contract_id,
            invoice_id=invoice_id,
            parts_billing_id=parts_billing_id,
            revenue_date=rdate,
            title=(title or '')[:300],
            revenue_type=rev_type,
            payment_method=_str(_cell(r, 'طريقة الدفع')) or 'كاش',
            amount=amount_ex,
            tax_amount=tax,
            total=total_incl,
            status=_map_status(_cell(r, 'Status', 'الحالة')),
            reference=_str(_cell(r, 'مرفقات'))[:500],
            notes=row_notes,
        )

        if existing_rev and sync_existing:
            revenue = existing_rev
            for key, val in fields.items():
                setattr(revenue, key, val)
            action = 'updated'
        else:
            revenue = Revenue(code=code, **fields)
            assign_organization(revenue)
            action = 'imported'

        if not dry_run:
            if action == 'imported':
                db.session.add(revenue)
            db.session.flush()

            if parts_billing_id and revenue.status in ('محصّل', 'محصل'):
                try:
                    link = apply_payment_to_source('parts_billing', parts_billing_id, total_incl)
                    revenue.parts_billing_id = link['parts_billing_id']
                    revenue.contract_id = link.get('contract_id') or revenue.contract_id
                    revenue.revenue_type = link.get('revenue_type') or revenue.revenue_type
                except ValueError:
                    pass

            if invoice_id and revenue.status in ('محصّل', 'محصل'):
                try:
                    link = apply_payment_to_source('invoice', invoice_id, total_incl)
                    revenue.invoice_id = link['invoice_id']
                    revenue.contract_id = link.get('contract_id') or revenue.contract_id
                    revenue.revenue_type = link.get('revenue_type') or revenue.revenue_type
                except ValueError:
                    pass

            existing_codes.add(code.upper())

        stats[action] += 1

    if not dry_run:
        repair_contract_payment_links(commit=False)
        for cid in contract_ids:
            sync_contract_invoice_status(cid)
        db.session.commit()

    stats['missing_samples'] = missing_samples
    stats['db_total'] = round(
        sum(_f(r.total) for r in tenant_query(Revenue).all()),
        2,
    ) if not dry_run else None
    missing_cn_amount = 0.0
    missing_cn_codes: list[str] = []
    for r in rows:
        raw_amount = _f(_cell(r, 'المبلغ'))
        if raw_amount <= 0:
            continue
        contract, customer, cn = _resolve_customer_contract(r)
        if not contract and not customer:
            missing_cn_amount = round(missing_cn_amount + raw_amount, 2)
            if cn and cn not in missing_cn_codes:
                missing_cn_codes.append(cn)
    stats['missing_amount'] = missing_cn_amount
    stats['missing_cn_count'] = len(missing_cn_codes)
    return stats


def main() -> int:
    parser = argparse.ArgumentParser(description='Import Jama revenues from Excel')
    parser.add_argument('xlsx', help='Path to revenues .xlsx')
    parser.add_argument('--slug', default='jama', help='Organization slug')
    parser.add_argument('--dry-run', action='store_true')
    parser.add_argument('--force', action='store_true', help='Import even if revenue code exists')
    parser.add_argument(
        '--sync',
        action='store_true',
        help='Update existing revenues from Excel and import missing rows',
    )
    parser.add_argument(
        '--import-all',
        action='store_true',
        help='Import every Excel row even if contract/customer not found in DB',
    )
    args = parser.parse_args()

    if not os.path.isfile(args.xlsx):
        print(f'ERROR: file not found: {args.xlsx}')
        return 1

    from flask import g
    from models import Organization

    with app.app_context():
        slug = (args.slug or 'jama').strip().lower()
        org = Organization.query.filter_by(slug=slug).first()
        if not org:
            print(f'ERROR: لا توجد مؤسسة slug={slug!r}')
            return 1
        g.organization = org
        g.organization_id = org.id
        print(f'Tenant: {org.name} ({org.slug})')
        print('File:', args.xlsx)
        result = import_revenues(
            args.xlsx,
            dry_run=args.dry_run,
            skip_existing=not (args.force or args.sync),
            sync_existing=args.sync,
            import_orphans=args.import_all,
        )
        print(result)
        if result.get('missing_samples'):
            print('Missing samples:')
            for line in result['missing_samples']:
                print(' ', line)
        if not args.dry_run:
            print('revenues in tenant:', tenant_query(Revenue).count())
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
