#!/usr/bin/env python3
"""استيراد حزمة جما (عملاء + فنيين + مصاعد + عقود) — بدون pandas (openpyxl فقط).

  set -a; source /etc/liftcore/platform.env; set +a
  cd ~/liftcore/elevator-app
  bash deploy/import_jama_tenant_bundle.sh
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / 'scripts'))


def _load_sheet_rows(path: str) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h or '').strip() for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(v is not None and str(v).strip() for v in r):
            continue
        d = {headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}
        out.append(d)
    return out


def _norm_phone(val) -> str:
    """طبّع جوال إلى +966XXXXXXXXX — يأخذ أول رقم إن وُجد أكثر من واحد."""
    raw = str(val or '').strip()
    if not raw or raw.lower() == 'nan':
        return ''
    # أكثر من رقم مفصول بفاصلة/شرطة/سطر
    first = re.split(r'[,;/|\n]+', raw)[0].strip()
    digits = re.sub(r'\D', '', first)
    if not digits:
        return ''
    if digits.startswith('966'):
        digits = digits[3:]
    if digits.startswith('0'):
        digits = digits[1:]
    if digits.startswith('5') and len(digits) >= 9:
        digits = digits[:9]
        return '+966' + digits
    if len(digits) >= 9:
        return '+' + digits[:15]
    return ''


def _secondary_phone(val) -> str:
    raw = str(val or '').strip()
    parts = [p.strip() for p in re.split(r'[,;/|\n]+', raw) if p.strip()]
    if len(parts) < 2:
        return ''
    return _norm_phone(parts[1])


def _bind_tenant(slug: str):
    from flask import g
    from models import Organization

    org = Organization.query.filter_by(slug=slug).first()
    if not org:
        print(f'ERROR: لا توجد مؤسسة slug={slug!r}')
        for o in Organization.query.order_by(Organization.id).all():
            print(f'  - {o.slug}')
        raise SystemExit(1)
    g.organization = org
    g.organization_id = org.id
    print(f'Tenant: {org.name} ({org.slug}) id={org.id}')
    return org


def _add(obj):
    from models import db
    from tenant_scope import assign_organization

    assign_organization(obj)
    db.session.add(obj)
    return obj


def import_customers(path: str, *, dry_run: bool = False) -> dict[str, int]:
    from import_real_data import _cell, _norm_city, _norm_id, _str
    from models import Customer, db

    rows = _load_sheet_rows(path)
    stats = {'rows': len(rows), 'added': 0, 'updated': 0, 'skipped': 0}

    def norm_code(code: str) -> str:
        s = _str(code)
        m = re.match(r'C-(\d+)', s, re.I)
        return f'C-{int(m.group(1)):04d}' if m else s.upper()

    existing = {norm_code(c.code): c for c in Customer.query.all() if c.code}

    for r in rows:
        code = norm_code(_cell(r, 'رقم العميل'))
        name = _str(_cell(r, 'اسم العميل')) or _str(_cell(r, 'اسم العميل | رقم العميل'))
        if '|' in name:
            name = name.split('|', 1)[0].strip()
        if not code or not name:
            stats['skipped'] += 1
            continue

        city = _norm_city(_cell(r, 'المدينة'))
        district = _str(_cell(r, 'الحي أو المنطقة', 'الحي'))
        address = _str(_cell(r, 'العنوان'))
        phone_raw = _cell(r, 'الجوال', 'الهاتف')
        phone = _norm_phone(phone_raw)
        phone2 = _secondary_phone(phone_raw)

        if code in existing:
            c = existing[code]
            changed = False
            for attr, val in (
                ('name', name),
                ('city', city),
                ('district', district),
                ('address', address or c.address),
                ('phone', phone or c.phone),
                ('phone2', phone2 or c.phone2),
            ):
                if val and getattr(c, attr) != val:
                    setattr(c, attr, val)
                    changed = True
            nid = _norm_id(_cell(r, 'رقم الهوية'))[:20]
            if nid and c.national_id != nid:
                c.national_id = nid
                changed = True
            email = _str(_cell(r, 'البريد الالكتروني', 'البريد الإلكتروني'))[:100]
            if email and c.email != email:
                c.email = email
                changed = True
            stats['updated' if changed else 'skipped'] += 1
            continue

        status = (_str(_cell(r, 'حالة العميل')) or 'نشط')[:20]
        c = Customer(
            code=code[:20],
            name=name[:200],
            city=city[:100],
            district=district[:100],
            address=address,
            phone=phone[:40],
            phone2=(phone2 or '')[:40],
            national_id=_norm_id(_cell(r, 'رقم الهوية'))[:20],
            email=_str(_cell(r, 'البريد الالكتروني', 'البريد الإلكتروني'))[:100],
            status=status,
            notes=_str(_cell(r, 'ملاحظات')),
            entity_type='فرد',
        )
        if dry_run:
            stats['added'] += 1
            existing[code] = c
            continue
        _add(c)
        existing[code] = c
        stats['added'] += 1

    if not dry_run:
        db.session.commit()
    return stats


def import_technicians(path: str, *, dry_run: bool = False) -> dict[str, int]:
    from import_real_data import _cell, _extract_tech, _str
    from models import Technician, db

    rows = _load_sheet_rows(path)
    stats = {'rows': len(rows), 'added': 0, 'updated': 0, 'skipped': 0}
    existing = {t.code.upper(): t for t in Technician.query.all() if t.code}

    def norm_status(raw: str) -> str:
        s = _str(raw).lower()
        if 'on duty' in s or 'نشط' in s or 'متاح' in s or 'مشغول' in s:
            return 'متاح'
        if 'off' in s or 'غير' in s:
            return 'غير نشط'
        return 'متاح'

    for r in rows:
        code = _extract_tech(_cell(r, 'Technical ID | رقم الفني', 'رقم واسم الفني', 'رقم الفني'))
        name = _str(_cell(r, 'Technical Name | اسم الفني', 'اسم الفني'))
        if not code or not name:
            stats['skipped'] += 1
            continue
        job = _str(_cell(r, 'Job Title | المسمى الوظيفي', 'المسمى الوظيفي'))
        if '|' in job:
            job = job.split('|')[-1].strip()
        status = norm_status(_cell(r, 'Status | الحالة', 'الحالة'))
        notes = _str(_cell(r, 'Notes | ملاحظات', 'ملاحظات'))

        if code in existing:
            t = existing[code]
            t.name = name
            t.job_title = job or t.job_title or 'فني مصاعد'
            t.status = status
            if notes:
                t.notes = notes
            stats['updated'] += 1
            continue

        t = Technician(
            code=code,
            name=name,
            job_title=job or 'فني مصاعد',
            status=status,
            city='مكة',
            emergency=True,
            notes=notes,
        )
        if dry_run:
            stats['added'] += 1
            existing[code] = t
            continue
        _add(t)
        existing[code] = t
        stats['added'] += 1

    if not dry_run:
        db.session.commit()
    return stats


def import_elevators(path: str, *, dry_run: bool = False) -> dict[str, int]:
    from import_elevators_xlsx import import_elevators as _imp
    from models import Elevator, db
    from tenant_scope import assign_organization

    orig_add = db.session.add

    def add_wrapped(obj):
        if isinstance(obj, Elevator) and getattr(obj, 'organization_id', None) is None:
            assign_organization(obj)
        return orig_add(obj)

    db.session.add = add_wrapped  # type: ignore[method-assign]
    try:
        return _imp(path, dry_run=dry_run)
    finally:
        db.session.add = orig_add  # type: ignore[method-assign]


def import_contracts(path: str, *, dry_run: bool = False) -> dict[str, int]:
    """عقود عبر openpyxl — بدون pandas."""
    from import_real_data import (
        _cell,
        _extract_all_el,
        _extract_cn,
        _f,
        _invoice_status,
        _norm_contract_status,
        _parse_date,
        _str,
    )
    from models import Contract, ContractElevator, Customer, Elevator, db
    from tenant_scope import assign_organization

    def norm_cn(code: str) -> str:
        s = _str(code)
        m = re.match(r'CN-(\d+)', s, re.I)
        return f'CN-{int(m.group(1)):05d}' if m else s.upper()

    def find_customer(name: str, cn_code: str | None) -> Customer | None:
        name = _str(name)
        if name:
            customer = Customer.query.filter_by(name=name).first()
            if customer:
                return customer
            norm = re.sub(r'\s+', ' ', name).strip()
            for c in Customer.query.all():
                if c.name and re.sub(r'\s+', ' ', c.name).strip() == norm:
                    return c
                if c.name and (norm in c.name or c.name in norm):
                    return c
        if cn_code:
            m = re.match(r'CN-(\d+)$', cn_code, re.I)
            if m:
                short = f'C-{int(m.group(1)):04d}'
                customer = Customer.query.filter_by(code=short).first()
                if customer:
                    return customer
        return None

    def link_elevators(contract: Contract, el_codes: list[str]) -> int:
        linked = 0
        for el_code in el_codes:
            if not el_code:
                continue
            elev = Elevator.query.filter_by(code=el_code).first()
            if not elev:
                continue
            exists = ContractElevator.query.filter_by(
                contract_id=contract.id, elevator_id=elev.id
            ).first()
            if exists:
                continue
            if dry_run:
                linked += 1
                continue
            ce = ContractElevator(contract_id=contract.id, elevator_id=elev.id)
            assign_organization(ce)
            db.session.add(ce)
            linked += 1
        return linked

    rows = _load_sheet_rows(path)
    stats = {
        'rows': len(rows),
        'added': 0,
        'updated': 0,
        'skipped': 0,
        'no_customer': 0,
        'linked_elevators': 0,
    }
    existing = {norm_cn(c.code): c for c in Contract.query.all() if c.code}

    for r in rows:
        code = _str(_cell(r, 'رقم العقد')) or (_extract_cn(_cell(r, 'اسم العميل ورقم العقد')) or '')
        code = norm_cn(code) if code else ''
        name = _str(_cell(r, 'العملاء'))
        annual = _f(_cell(r, 'قيمة العقد'))
        start = _parse_date(_cell(r, 'تاريخ بداية العقد'))
        end = _parse_date(_cell(r, 'تاريخ انتهاء العقد'))
        el_codes = _extract_all_el(_cell(r, 'رقم المصعد', 'اسم العميل ورقم العقد'))

        if not code or not start or not end:
            stats['skipped'] += 1
            continue
        if (not name or name.lower() == 'nan') and annual <= 0 and code not in existing:
            stats['skipped'] += 1
            continue

        customer = find_customer(name, code)
        if not customer:
            stats['no_customer'] += 1
            print(f'  [تخطي] {code}: لا عميل لـ «{name}»')
            continue

        paid = _f(_cell(r, 'المبلغ المسدد'))
        val = annual
        tax = round(val * 0.15, 2)
        payload = dict(
            customer_id=customer.id,
            contract_type=(
                'عقد صيانة'
                if _str(_cell(r, 'نوع العقد')) == 'صيانة'
                else _str(_cell(r, 'نوع العقد')) or 'عقد صيانة'
            ),
            start_date=start,
            end_date=end,
            duration_months=max(0, (end.year - start.year) * 12 + end.month - start.month),
            maint_frequency=_str(_cell(r, 'برنامج الصيانة')) or 'سنوي',
            visits_per_month=1,
            value=val,
            tax_pct=15,
            tax_amount=tax,
            total=round(val + tax, 2),
            payment_terms='دفعة واحدة',
            invoice_status=_invoice_status(val, paid),
            status=_norm_contract_status(_cell(r, 'حالة العقد')),
            reminder_date=end - timedelta(days=30),
            city=customer.city or _str(_cell(r, 'المنطقة')),
            district=customer.district or '',
            address=customer.address or _str(_cell(r, 'العنوان')),
            notes=_str(_cell(r, 'ملاحظات')),
        )

        if code in existing:
            c = existing[code]
            for key, val_item in payload.items():
                setattr(c, key, val_item)
            stats['updated'] += 1
            contract = c
        else:
            if dry_run:
                stats['added'] += 1
                stats['linked_elevators'] += sum(
                    1 for el in el_codes if el and Elevator.query.filter_by(code=el).first()
                )
                continue
            c = Contract(code=code, **payload)
            assign_organization(c)
            db.session.add(c)
            db.session.flush()
            existing[code] = c
            contract = c
            stats['added'] += 1

        if contract.id:
            stats['linked_elevators'] += link_elevators(contract, el_codes)

    if not dry_run:
        db.session.commit()
    return stats


def main() -> int:
    parser = argparse.ArgumentParser(description='Import Jama bundle into a tenant')
    parser.add_argument('--slug', default='jama')
    parser.add_argument('--clients', required=True)
    parser.add_argument('--technicians', required=True)
    parser.add_argument('--elevators', required=True)
    parser.add_argument('--contracts', required=True)
    parser.add_argument('--dry-run', action='store_true')
    parser.add_argument('--skip-geocode', action='store_true', default=True)
    parser.add_argument('--geocode', action='store_true', help='Geocode customer addresses')
    args = parser.parse_args()

    for label, path in (
        ('clients', args.clients),
        ('technicians', args.technicians),
        ('elevators', args.elevators),
        ('contracts', args.contracts),
    ):
        if not os.path.isfile(path):
            print(f'ERROR: {label} not found: {path}')
            return 1

    from app import app, db
    from models import Contract, Customer, Elevator, Technician

    with app.app_context():
        org = _bind_tenant(args.slug.strip().lower())
        print('Database:', (app.config.get('SQLALCHEMY_DATABASE_URI') or '')[:60], '...')

        print('\n==> [1/4] العملاء')
        print(import_customers(args.clients, dry_run=args.dry_run))

        print('\n==> [2/4] الفنيين')
        print(import_technicians(args.technicians, dry_run=args.dry_run))

        print('\n==> [3/4] المصاعد')
        print(import_elevators(args.elevators, dry_run=args.dry_run))

        print('\n==> [4/4] العقود')
        print(import_contracts(args.contracts, dry_run=args.dry_run))

        if args.geocode and not args.dry_run:
            print('\n==> Geocode')
            from client_address_import import geocode_customers_missing
            print(geocode_customers_missing(db_session=db.session))

        if not args.dry_run:
            print('\n=== الملخص (جما) ===')
            print('  عملاء:', Customer.query.filter_by(organization_id=org.id).count())
            print('  فنيون:', Technician.query.filter_by(organization_id=org.id).count())
            print('  مصاعد:', Elevator.query.filter_by(organization_id=org.id).count())
            print('  عقود:', Contract.query.filter_by(organization_id=org.id).count())
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
