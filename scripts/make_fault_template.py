#!/usr/bin/env python3
"""توليد قالب إكسل لبيانات الأعطال (للتعبئة اليدوية ثم الاستيراد لاحقاً).

  python scripts/make_fault_template.py
الناتج: deploy/data/fault_import_template.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, 'deploy', 'data', 'fault_import_template.xlsx')

# (العنوان، العرض، تلميح، إلزامي)
COLUMNS = [
    ('كود المصعد', 16, 'كود المصعد كما في البرنامج، مثل EL-00001 (إلزامي للربط)', True),
    ('تاريخ البلاغ', 14, 'صيغة YYYY-MM-DD مثل 2026-08-08 (إلزامي)', True),
    ('وقت البلاغ', 12, 'صيغة HH:MM مثل 14:30 (اختياري)', False),
    ('نوع العطل', 20, 'مثل: توقف مفاجئ، عطل باب، صوت غير طبيعي...', False),
    ('الأولوية', 12, 'عادية / عاجلة / حرجة', False),
    ('الحالة', 16, 'مفتوح / قيد المعالجة / انتظار قطع / تم الاصلاح / مغلق', False),
    ('وصف البلاغ', 34, 'ما ذكره العميل عند الإبلاغ', False),
    ('اسم المُبلِّغ', 18, 'اسم الشخص الذي أبلغ عن العطل', False),
    ('هاتف المُبلِّغ', 16, 'رقم جوال المُبلِّغ', False),
    ('الفني', 18, 'اسم الفني أو كوده (اختياري)', False),
    ('التشخيص الفني', 34, 'سبب العطل بعد الفحص', False),
    ('الإجراء / الحل', 34, 'ما تم عمله لإصلاح العطل', False),
    ('يحتاج قطع غيار', 14, 'نعم / لا', False),
    ('كود الزيارة المرتبطة', 18, 'كود زيارة الصيانة إن وُجد، مثل VI-00001 (اختياري)', False),
    ('كود العقد', 14, 'مثل CN-00001 (اختياري)', False),
    ('تاريخ الإصلاح', 14, 'YYYY-MM-DD إن تم الإصلاح (اختياري)', False),
    ('ملاحظات', 30, 'أي ملاحظات إضافية', False),
]

SAMPLE = [
    'EL-00001', '2026-08-08', '14:30', 'توقف مفاجئ', 'عاجلة', 'تم الاصلاح',
    'المصعد واقف بين الدورين', 'أبو محمد', '0555555555', 'خالد',
    'عطل في لوحة التحكم', 'استبدال ريلاي وإعادة التشغيل', 'نعم',
    'VI-00012', 'CN-00003', '2026-08-08', 'تم التواصل مع العميل',
]


def main() -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = 'الأعطال'
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill('solid', fgColor='1A3A5C')
    req_fill = PatternFill('solid', fgColor='C0392B')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin = Side(style='thin', color='B0B0B0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for idx, (title, width, hint, required) in enumerate(COLUMNS, start=1):
        col = get_column_letter(idx)
        ws.column_dimensions[col].width = width
        cell = ws.cell(row=1, column=idx, value=title + (' *' if required else ''))
        cell.fill = req_fill if required else header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        cell.comment = Comment(hint, 'LiftCore')

    for idx, value in enumerate(SAMPLE, start=1):
        cell = ws.cell(row=2, column=idx, value=value)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = border
        cell.font = Font(color='888888', italic=True)

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 34

    # قوائم منسدلة للتحقق
    def add_list(col_letter: str, options: str):
        dv = DataValidation(type='list', formula1=f'"{options}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f'{col_letter}3:{col_letter}1000')

    add_list('E', 'عادية,عاجلة,حرجة')
    add_list('F', 'مفتوح,قيد المعالجة,انتظار قطع,تم الاصلاح,مغلق')
    add_list('M', 'نعم,لا')

    # ورقة تعليمات
    info = wb.create_sheet('تعليمات')
    info.sheet_view.rightToLeft = True
    info.column_dimensions['A'].width = 90
    lines = [
        ('تعليمات تعبئة بيانات الأعطال', True),
        ('', False),
        ('• الأعمدة المميّزة بنجمة (*) وبلون أحمر إلزامية.', False),
        ('• «كود المصعد» يجب أن يطابق كود المصعد في البرنامج (مثل EL-00001) حتى يُربط العطل بالمصعد الصحيح.', False),
        ('• التواريخ بصيغة YYYY-MM-DD (سنة-شهر-يوم)، والوقت بصيغة HH:MM بنظام 24 ساعة.', False),
        ('• «الأولوية»: عادية / عاجلة / حرجة.', False),
        ('• «الحالة»: مفتوح / قيد المعالجة / انتظار قطع / تم الاصلاح / مغلق.', False),
        ('• «يحتاج قطع غيار»: نعم / لا.', False),
        ('• الصف الثاني الرمادي مثال توضيحي — احذفه قبل الإرسال أو اكتب فوقه.', False),
        ('• أضف صفاً واحداً لكل عطل. لا تغيّر ترتيب الأعمدة أو عناوينها.', False),
        ('• الأعمدة الاختيارية اتركها فارغة إن لم تتوفر.', False),
    ]
    for i, (text, is_title) in enumerate(lines, start=1):
        c = info.cell(row=i, column=1, value=text)
        c.alignment = Alignment(horizontal='right', wrap_text=True)
        if is_title:
            c.font = Font(bold=True, size=14, color='1A3A5C')

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print('تم إنشاء القالب:', OUT)
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
