#!/usr/bin/env python3
"""استيراد عملاء من Excel إلى مؤسسة (واجهة /clients) عبر API الداخلي.

  # على السيرفر — جما:
  set -a; source /etc/liftcore/platform.env; set +a
  cd ~/liftcore/elevator-app
  .venv/bin/python scripts/import_clients_ui_xlsx.py \\
    --slug jama \\
    --xlsx static/templates/clients_import_jama_ready.xlsx

  # أو ملف Downloads بعد رفعه:
  .venv/bin/python scripts/import_clients_ui_xlsx.py --slug jama --xlsx /tmp/clients_template_fixed.xlsx
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def _load_rows(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h or '').strip() for h in rows[0]]
    out = []
    for r in rows[1:]:
        d = {headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}
        if any(v not in (None, '') for v in d.values()):
            out.append(d)
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description='Import clients Excel into a tenant')
    parser.add_argument('--slug', required=True, help='Organization slug (jama, default, …)')
    parser.add_argument('--xlsx', required=True, help='Path to .xlsx')
    parser.add_argument('--limit', type=int, default=0, help='Import only first N rows (0=all)')
    args = parser.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.is_file():
        print(f'ERROR: file not found: {xlsx}')
        return 1

    from flask import g

    from app import app
    from client_bulk_import import import_customer_rows
    from models import Customer, Organization

    rows = _load_rows(xlsx)
    if args.limit and args.limit > 0:
        rows = rows[: args.limit]
    print(f'File: {xlsx} — rows: {len(rows)}')

    with app.app_context():
        slug = args.slug.strip().lower()
        org = Organization.query.filter_by(slug=slug).first()
        if not org:
            print(f'ERROR: لا توجد مؤسسة slug={slug!r}')
            for o in Organization.query.order_by(Organization.id).all():
                print(f'  - {o.slug}')
            return 1
        g.organization = org
        g.organization_id = org.id
        before = Customer.query.filter_by(organization_id=org.id).count()
        print(f'Tenant: {org.name} ({org.slug}) — قبل: {before} عميل')
        result = import_customer_rows(rows)
        after = Customer.query.filter_by(organization_id=org.id).count()
        print('=== النتيجة ===')
        print(f'  imported: {result["imported"]}')
        print(f'  failed:   {result["failed"]}')
        print(f'  after:    {after} عميل')
        for err in result.get('errors') or []:
            print(f'  • صف {err["row"]}: {err["error"]}')
        return 0 if result['imported'] or not rows else 1


if __name__ == '__main__':
    raise SystemExit(main())
