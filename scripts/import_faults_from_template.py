#!/usr/bin/env python3
"""استيراد الأعطال من قالب الإكسل (fault_import_template.xlsx) إلى قاعدة بيانات المستأجر.

يقرأ الأعمدة بالاسم (لا بالترتيب)، ويربط كل عطل بالمصعد عبر «كود المصعد».
يتخطى الصف إذا كان عطلاً مطابقاً (نفس المصعد + نفس تاريخ/وقت البلاغ) موجوداً مسبقاً.

  cd ~/liftcore/jama-elevator-app
  source .venv/bin/activate
  set -a; source /etc/liftcore/platform.env; set +a   # أو export DATABASE_URL=...
  python scripts/import_faults_from_template.py deploy/data/fault_import_template_filled.xlsx --slug jama --dry-run
  python scripts/import_faults_from_template.py deploy/data/fault_import_template_filled.xlsx --slug jama --yes
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import datetime

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit('pip install openpyxl') from exc


HEADER_MAP = {
    'كود المصعد': 'elevator',
    'تاريخ البلاغ': 'date',
    'وقت البلاغ': 'time',
    'نوع العطل': 'fault_type',
    'الأولوية': 'priority',
    'الحالة': 'status',
    'وصف البلاغ': 'client_report',
    'اسم المبلغ': 'reporter_name',
    'هاتف المبلغ': 'reporter_phone',
    'الفني': 'technician',
    'التشخيص الفني': 'tech_notes',
    'الإجراء / الحل': 'resolution',
    'الاجراء / الحل': 'resolution',
    'يحتاج قطع غيار': 'needs_parts',
    'كود الزيارة المرتبطة': 'visit_code',
    'كود العقد': 'contract_code',
    'تاريخ الإصلاح': 'resolved_date',
    'تاريخ الاصلاح': 'resolved_date',
    'ملاحظات': 'notes',
}

VALID_STATUS = {'مفتوح', 'قيد المعالجة', 'انتظار قطع', 'تم الاصلاح', 'مغلق'}


def _s(val) -> str:
    if val is None:
        return ''
    s = str(val).strip()
    return '' if s.lower() == 'nan' else s


def _norm_header(h: str) -> str:
    h = _s(h).replace('*', '').replace('ـ', '').strip()
    h = re.sub(r'\s+', ' ', h)
    # توحيد الهمزات الشائعة في «المبلِّغ»
    h = h.replace('المُبلِّغ', 'المبلغ').replace('المبلِّغ', 'المبلغ').replace('المُبلغ', 'المبلغ')
    return h


def _parse_dt(date_s: str, time_s: str):
    date_s, time_s = _s(date_s), _s(time_s)
    d = None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            d = datetime.strptime(date_s[:10], fmt)
            break
        except ValueError:
            continue
    if d is None:
        return None
    if time_s:
        for fmt in ('%H:%M', '%H:%M:%S', '%I:%M %p'):
            try:
                t = datetime.strptime(time_s, fmt)
                return d.replace(hour=t.hour, minute=t.minute)
            except ValueError:
                continue
    return d


def _map_status(raw: str) -> str:
    s = _s(raw)
    if s in VALID_STATUS:
        return s
    if 'اصلاح' in s or 'مكتمل' in s or 'محلول' in s:
        return 'تم الاصلاح'
    if 'انتظار' in s and 'قطع' in s:
        return 'انتظار قطع'
    if 'قيد' in s or 'جار' in s:
        return 'قيد المعالجة'
    if 'مغلق' in s or 'ملغ' in s:
        return 'مغلق'
    return 'مفتوح'


def _norm_el(code: str) -> str:
    m = re.search(r'EL-\s*(\d+)', _s(code), re.I)
    return f'EL-{int(m.group(1)):04d}' if m else _s(code).upper()


def load_template_rows(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['الأعطال'] if 'الأعطال' in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = rows[0]
    col_to_key = {}
    for i, h in enumerate(header):
        key = HEADER_MAP.get(_norm_header(h))
        if key:
            col_to_key[i] = key
    out = []
    for r in rows[1:]:
        if not any(_s(v) for v in r):
            continue
        rec = {v: '' for v in HEADER_MAP.values()}
        for i, key in col_to_key.items():
            if i < len(r):
                rec[key] = _s(r[i])
        out.append(rec)
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description='استيراد الأعطال من قالب الإكسل')
    parser.add_argument('xlsx', help='مسار ملف الإكسل')
    parser.add_argument('--slug', default='jama')
    parser.add_argument('--dry-run', action='store_true')
    parser.add_argument('--yes', action='store_true')
    args = parser.parse_args()

    if not os.path.isfile(args.xlsx):
        print(f'ERROR: file not found: {args.xlsx}')
        return 1
    if not args.dry_run and not args.yes:
        print('أضف --yes للتأكيد أو --dry-run للمعاينة')
        return 2

    import json

    from flask import g
    from app import app, db, next_code
    from fault_report import empty_report
    from models import Elevator, Fault, MaintenanceVisit, Organization, Technician
    from tenant_scope import assign_organization

    try:
        from entity_links import link_fault_to_visit
    except Exception:
        link_fault_to_visit = None

    records = load_template_rows(args.xlsx)

    with app.app_context():
        slug = (args.slug or 'jama').strip().lower()
        org = Organization.query.filter_by(slug=slug).first()
        if not org:
            print(f'ERROR: لا توجد مؤسسة slug={slug!r}')
            return 1
        g.organization = org
        g.organization_id = org.id
        print(f'Tenant: {org.name} ({org.slug}) — صفوف الملف: {len(records)}')

        # فهارس البحث
        el_idx: dict[str, object] = {}
        for e in Elevator.query.filter_by(organization_id=org.id).all():
            if e.code:
                el_idx[e.code.upper()] = e
                el_idx[_norm_el(e.code)] = e
        tech_by_name: dict[str, object] = {}
        tech_by_code: dict[str, object] = {}
        for t in Technician.query.filter_by(organization_id=org.id).all():
            if t.name:
                tech_by_name[t.name.strip()] = t
            if getattr(t, 'code', None):
                tech_by_code[t.code.upper()] = t
        visits_by_code = {
            v.code.upper(): v
            for v in MaintenanceVisit.query.filter_by(organization_id=org.id).all()
            if v.code
        }
        existing = {
            (f.elevator_id, f.reported_at)
            for f in Fault.query.filter_by(organization_id=org.id).all()
        }

        stats = {'imported': 0, 'skipped_existing': 0, 'missing_elevator': 0, 'errors': 0}
        missing_samples: list[str] = []

        for rec in records:
            el_code = _norm_el(rec['elevator'])
            reported_at = _parse_dt(rec['date'], rec['time'])
            if not el_code or not reported_at:
                stats['errors'] += 1
                continue
            elevator = el_idx.get(el_code) or el_idx.get(rec['elevator'].upper())
            if not elevator:
                stats['missing_elevator'] += 1
                if len(missing_samples) < 15:
                    missing_samples.append(rec['elevator'])
                continue
            if (elevator.id, reported_at) in existing:
                stats['skipped_existing'] += 1
                continue

            tech = None
            tname = rec['technician'].strip()
            if tname:
                tech = tech_by_name.get(tname)
                if not tech:
                    m = re.search(r'Tech-\s*(\d+)', tname, re.I)
                    if m:
                        tech = tech_by_code.get(f'TECH-{int(m.group(1)):03d}')
            status = _map_status(rec['status'])
            resolved_at = _parse_dt(rec['resolved_date'], '') or (
                reported_at if status == 'تم الاصلاح' else None
            )
            priority = rec['priority'].strip() or 'عادية'
            needs_parts = rec['needs_parts'].strip() in ('نعم', 'yes', 'true', '1')
            notes = rec['notes'].strip()
            if rec['contract_code']:
                meta = f"عقد: {rec['contract_code']}"
                notes = f'{notes}\n{meta}' if notes else meta

            if args.dry_run:
                stats['imported'] += 1
                continue

            has_time = bool(rec['time'].strip())
            report = empty_report()
            report['meta'].update({
                'visit_date': reported_at.strftime('%Y-%m-%d'),
                'arrival_time': reported_at.strftime('%H:%M') if has_time else '',
                'client_description': rec['client_report'],
                'fault_types': [rec['fault_type']] if rec['fault_type'] else [],
                'diagnosis': rec['tech_notes'],
                'action_taken': rec['resolution'],
                'visit_outcome': 'solved' if status == 'تم الاصلاح' else '',
                'final_notes': rec['notes'],
                'contract_type': 'عقد صيانة نشط' if rec['contract_code'] else 'بدون عقد',
            })
            if resolved_at:
                report['meta']['end_time'] = resolved_at.strftime('%H:%M') if has_time else ''

            fault = Fault(
                code=next_code(Fault, 'FA-', digits=5),
                elevator_id=elevator.id,
                technician_id=tech.id if tech else None,
                fault_type=rec['fault_type'] or 'عطل',
                description=rec['client_report'] or rec['resolution'] or rec['fault_type'] or 'عطل',
                client_report=rec['client_report'] or None,
                reporter_name=rec['reporter_name'] or None,
                reporter_phone=rec['reporter_phone'] or None,
                tech_notes=rec['tech_notes'] or None,
                needs_parts=needs_parts,
                priority=priority,
                reported_at=reported_at,
                status=status,
                resolution=rec['resolution'] or None,
                resolved_at=resolved_at,
                notes=notes or None,
                report_json=json.dumps(report, ensure_ascii=False),
            )
            assign_organization(fault)
            db.session.add(fault)
            db.session.flush()

            vi = rec['visit_code'].strip().upper()
            if vi and link_fault_to_visit:
                m = re.search(r'VI-\s*(\d+)', vi, re.I)
                key = f'VI-{int(m.group(1)):05d}' if m else vi
                visit = visits_by_code.get(key) or visits_by_code.get(vi)
                if visit:
                    try:
                        link_fault_to_visit(fault, visit)
                    except Exception:
                        pass

            existing.add((elevator.id, reported_at))
            stats['imported'] += 1

        if not args.dry_run:
            db.session.commit()

        print('=== النتيجة ===')
        print(f"  {'سيُستورد' if args.dry_run else 'مستورد'}: {stats['imported']}")
        print(f"  متخطى (موجود مسبقاً): {stats['skipped_existing']}")
        print(f"  مصعد غير موجود: {stats['missing_elevator']}")
        print(f"  أخطاء (بيانات ناقصة): {stats['errors']}")
        if missing_samples:
            print('  عيّنة أكواد مصاعد غير موجودة:', ', '.join(missing_samples))
        if not args.dry_run:
            print('  إجمالي أعطال المستأجر الآن:',
                  Fault.query.filter_by(organization_id=org.id).count())
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
