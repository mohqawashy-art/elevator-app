#!/usr/bin/env python3
"""إصلاح clients_template.xlsx في Downloads — أعمدة بمسافات + بيانات جما الصحيحة."""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parents[1]
JAMA = ROOT / 'deploy' / 'data' / 'jama_import' / 'العملاء 1_7_2026.xlsx'
OUT = Path(r'c:\Users\HOME\Downloads\clients_template_fixed.xlsx')
OUT_FULL = ROOT / 'static' / 'templates' / 'clients_import_jama_ready.xlsx'
BROKEN = Path(r'c:\Users\HOME\Downloads\clients_template.xlsx')
# إن كان الملف الأصلي مقفولاً نقرأ منه إن أمكن، ونكتب دائماً على الاسم الجديد
OUT_ALT = Path(r'c:\Users\HOME\Downloads\clients_ready_import.xlsx')

HEADERS = [
    'الاسم (عربي)',
    'المدينة',
    'الحي',
    'العنوان',
    'رقم الهاتف',
    'البريد الإلكتروني',
    'اسم المسؤول',
    'نوع المتعاقد',
    'رقم هوية المتعاقد',
    'رقم السجل التجاري',
]


def _phone9(raw) -> str:
    digits = re.sub(r'\D', '', str(raw or ''))
    if digits.startswith('966'):
        digits = digits[3:]
    if digits.startswith('0'):
        digits = digits[1:]
    return digits


def load_jama_rows() -> list[dict]:
    wb = load_workbook(JAMA, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h or '').strip() for h in rows[0]]
    wb.close()
    out = []
    for r in rows[1:]:
        d = dict(zip(headers, r))
        name = str(d.get('اسم العميل') or '').strip()
        if not name:
            continue
        phone = _phone9(d.get('الجوال'))
        email = d.get('البريد الالكتروني')
        out.append({
            'name': name,
            'city': str(d.get('المدينة') or '').strip(),
            'district': str(d.get('الحي أو المنطقة') or '').strip(),
            'address': str(d.get('العنوان') or '').strip(),
            'phone': phone,
            'email': str(email).strip() if email else '',
            'contact': name,
            'entity': 'فرد',
            'nid': str(d.get('رقم الهوية') or '').strip(),
            'cr': '',
            'phone_key': phone[-9:] if phone else '',
        })
    return out


def phones_from_broken() -> list[str]:
    if not BROKEN.is_file():
        return []
    wb = load_workbook(BROKEN, data_only=True)
    ws = wb.active
    keys = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        key = _phone9(row[4])[-9:]
        if key:
            keys.append(key)
    wb.close()
    return keys


def write_xlsx(path: Path, records: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Clients'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2A7FFF')
    thin = Border(
        left=Side(style='thin', color='D0D7E2'),
        right=Side(style='thin', color='D0D7E2'),
        top=Side(style='thin', color='D0D7E2'),
        bottom=Side(style='thin', color='D0D7E2'),
    )
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin

    for i, rec in enumerate(records, 2):
        vals = [
            rec['name'],
            rec['city'],
            rec['district'],
            rec['address'],
            str(rec['phone']),
            rec['email'],
            rec['contact'],
            rec['entity'],
            rec['nid'],
            rec['cr'],
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(i, col, '' if val in (None, 'None') else val)
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.border = thin
            if col == 5:
                cell.number_format = '@'
                cell.value = str(rec['phone'])

    widths = [28, 16, 14, 48, 14, 24, 22, 12, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.auto_filter.ref = f'A1:J{len(records) + 1}'
    ws.freeze_panes = 'A2'
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f'[OK] {path} — {len(records)} clients')


def main() -> int:
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    all_rows = load_jama_rows()
    by_phone = {r['phone_key']: r for r in all_rows if r['phone_key']}
    want = phones_from_broken()
    selected = [by_phone[k] for k in want if k in by_phone]
    if len(selected) < 10:
        selected = all_rows[:15]
    write_xlsx(OUT, selected)
    try:
        write_xlsx(BROKEN, selected)
    except PermissionError:
        print('[WARN] تعذر الكتابة على clients_template.xlsx (مفتوح؟) — استخدم الملف الجديد')
    write_xlsx(OUT_ALT, selected)
    write_xlsx(OUT_FULL, all_rows)
    print('sample:', selected[0]['name'], '|', selected[0]['phone'], '|', selected[0]['city'])
    print('ارفع هذا الملف:', OUT)
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
