"""تحديث عناوين العملاء من Excel (مطابقة برقم العميل) + إحداثيات للخريطة."""

from __future__ import annotations

import io
import os
import re
from typing import BinaryIO

try:
    import openpyxl
except ImportError as exc:
    raise ImportError('pip install openpyxl') from exc

from geocode import geocode_customer
from models import Customer


def _str(val) -> str:
    if val is None:
        return ''
    s = str(val).strip()
    return '' if s.lower() == 'nan' else s


def _extract_code(text: str) -> str | None:
    m = re.search(r'C-\d+', _str(text), re.I)
    return m.group(0).upper() if m else None


def normalize_client_code(code: str) -> str:
    m = re.match(r'C-(\d+)', _str(code), re.I)
    if m:
        return f'C-{int(m.group(1)):04d}'
    return _str(code).upper()


def _row_dict(headers: list[str], values: tuple) -> dict[str, str]:
    row = {}
    for i, h in enumerate(headers):
        key = _str(h)
        if not key:
            continue
        row[key] = _str(values[i]) if i < len(values) else ''
    return row


def _cell(row: dict[str, str], *names: str) -> str:
    for name in names:
        if name in row and row[name]:
            return row[name]
    return ''


def _name_from_row(row: dict[str, str]) -> str:
    combined = _cell(row, 'اسم العميل | رقم العميل', 'اسم العميل', 'name', 'الاسم')
    if '|' in combined:
        return combined.split('|', 1)[0].strip()
    return combined


def _find_header_row(rows: list[tuple]) -> int:
    for i, values in enumerate(rows):
        text = ' '.join(_str(v) for v in values)
        if 'رقم العميل' in text or 'C-' in text:
            if 'العنوان' in text or 'المدينة' in text:
                return i
    return 0


def _parse_workbook(wb) -> list[dict[str, str]]:
    ws = wb.active
    raw = [tuple(r) for r in ws.iter_rows(values_only=True)]
    if not raw:
        return []
    hi = _find_header_row(raw)
    headers = [_str(h) for h in raw[hi]]
    out = []
    for values in raw[hi + 1:]:
        if not any(_str(v) for v in values):
            continue
        row = _row_dict(headers, values)
        code = _cell(row, 'رقم العميل', 'كود العميل', 'customer_code', 'Code')
        if not code:
            code = _extract_code(_cell(row, 'اسم العميل | رقم العميل', 'اسم العميل', 'name')) or ''
        if not code:
            for v in values:
                code = _extract_code(_str(v))
                if code:
                    break
        if not code:
            continue
        row['_code'] = normalize_client_code(code)
        row['_name'] = _name_from_row(row)
        out.append(row)
    return out


def load_rows_from_path(path: str) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return _parse_workbook(wb)
    finally:
        wb.close()


def load_rows_from_stream(stream: BinaryIO) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(stream, read_only=True, data_only=True)
    try:
        return _parse_workbook(wb)
    finally:
        wb.close()


def _build_customer_index(customers: list[Customer]) -> tuple[dict[str, Customer], dict[str, Customer]]:
    by_code: dict[str, Customer] = {}
    by_name: dict[str, Customer] = {}
    for customer in customers:
        if customer.code:
            by_code[normalize_client_code(customer.code)] = customer
            by_code[customer.code.upper()] = customer
        name = _str(customer.name)
        if name:
            by_name[name] = customer
    return by_code, by_name


def _display_address(city: str, district: str, geo_address: str) -> str:
    """عنوان عربي قصير للواجهة — الخريطة تعتمد lat/lng وليس هذا الحقل."""
    if district and city:
        return f'{district}، {city}'
    if district:
        return district
    if city:
        return city
    if geo_address:
        return geo_address.split(',')[0].strip()
    return ''


def _lookup_customer(
    row: dict[str, str],
    by_code: dict[str, Customer],
    by_name: dict[str, Customer],
) -> Customer | None:
    code = row.get('_code', '')
    customer = by_code.get(code) or by_code.get(normalize_client_code(code))
    if customer:
        return customer
    name = row.get('_name', '')
    if name:
        return by_name.get(name)
    return None


def import_client_addresses(
    rows: list[dict[str, str]],
    *,
    dry_run: bool = False,
    no_geocode: bool = False,
    force_geocode: bool = False,
    geocode_delay: float = 0.45,
    db_session=None,
) -> dict:
    """تحديث العناوين وإرجاع ملخص."""
    updated = missing = geocoded = geo_fail = skipped = 0
    missing_codes: list[str] = []

    customers = Customer.query.all()
    by_code, by_name = _build_customer_index(customers)

    for row in rows:
        code = row.get('_code', '')
        customer = _lookup_customer(row, by_code, by_name)
        if not customer:
            missing += 1
            if len(missing_codes) < 25:
                missing_codes.append(code or row.get('_name', '?'))
            continue

        city = _cell(row, 'المدينة', 'city')
        district = _cell(row, 'الحي أو المنطقة', 'الحي', 'district')
        geo_query = _cell(row, 'العنوان', 'address', 'العنوان التفصيلي')
        display_addr = _display_address(city, district, geo_query)

        if not geo_query and not city and not district:
            skipped += 1
            continue

        changed = False
        if city and customer.city != city:
            customer.city = city
            changed = True
        if district and customer.district != district:
            customer.district = district
            changed = True
        if display_addr and customer.address != display_addr:
            customer.address = display_addr
            changed = True

        need_geo = not no_geocode and (
            force_geocode
            or changed
            or not (customer.lat and customer.lng)
        )
        if need_geo:
            has_gps = False
            if customer.lat and customer.lng and not force_geocode:
                try:
                    float(customer.lat)
                    float(customer.lng)
                    has_gps = True
                except (TypeError, ValueError):
                    pass
            if not has_gps or force_geocode:
                if dry_run:
                    changed = True
                elif geocode_customer(
                    customer,
                    delay=geocode_delay,
                    query_address=geo_query or None,
                    force=force_geocode,
                ):
                    geocoded += 1
                    changed = True
                else:
                    geo_fail += 1
                    changed = True

        if changed:
            updated += 1

    if not dry_run and db_session is not None:
        db_session.commit()

    return {
        'rows': len(rows),
        'updated': updated,
        'missing': missing,
        'skipped': skipped,
        'geocoded': geocoded,
        'geo_fail': geo_fail,
        'missing_codes': missing_codes,
        'dry_run': dry_run,
    }


def geocode_customers_missing(
    *,
    dry_run: bool = False,
    force: bool = False,
    geocode_delay: float = 0.35,
    db_session=None,
) -> dict:
    """تحديد مواقع الخريطة (lat/lng) للعملاء الذين لديهم عنوان/مدينة بدون GPS."""
    geocoded = geo_fail = skipped = 0
    for customer in Customer.query.all():
        has_gps = False
        if customer.lat and customer.lng and not force:
            try:
                float(customer.lat)
                float(customer.lng)
                has_gps = True
            except (TypeError, ValueError):
                pass
        if has_gps:
            continue
        if not (customer.address or customer.city or customer.district):
            skipped += 1
            continue
        if dry_run:
            geocoded += 1
            continue
        if geocode_customer(customer, delay=geocode_delay, force=force):
            geocoded += 1
        else:
            geo_fail += 1

    if not dry_run and db_session is not None:
        db_session.commit()

    return {
        'geocoded': geocoded,
        'geo_fail': geo_fail,
        'skipped': skipped,
        'dry_run': dry_run,
    }


def import_client_addresses_file(
    source,
    *,
    dry_run: bool = False,
    no_geocode: bool = False,
    force_geocode: bool = False,
    geocode_delay: float = 0.45,
    db_session=None,
) -> dict:
    if isinstance(source, (str, os.PathLike)):
        rows = load_rows_from_path(os.fspath(source))
    elif isinstance(source, (bytes, bytearray)):
        rows = load_rows_from_stream(io.BytesIO(source))
    elif hasattr(source, 'read'):
        data = source.read()
        if isinstance(data, str):
            data = data.encode('utf-8')
        rows = load_rows_from_stream(io.BytesIO(data))
    else:
        raise TypeError('source must be a path, bytes, or a readable file object')
    return import_client_addresses(
        rows,
        dry_run=dry_run,
        no_geocode=no_geocode,
        force_geocode=force_geocode,
        geocode_delay=geocode_delay,
        db_session=db_session,
    )
