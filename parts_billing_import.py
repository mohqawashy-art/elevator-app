"""استيراد بيان تركيب قطع الغيار من Excel."""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import BinaryIO

try:
    import openpyxl
except ImportError as exc:
    raise ImportError('pip install openpyxl') from exc

from entity_links import normalize_parts_status, resolve_parts_links
from models import Contract, PartsBilling

OP_NOTE_PREFIX = 'رقم العملية:'

COL_ALIASES = {
    'title': ('title', 'رقم العقد و اسم العميل'),
    'op': ('رقم العمليه', 'رقم العملية', 'operation'),
    'contract': ('العقود', 'رقم العقد', 'contract'),
    'date': ('التاريخ', 'billing_date', 'تاريخ'),
    'description': ('بيان قطع الغيار', 'description', 'البيان'),
    'cost': ('سعر التكلفة', 'cost', 'التكلفة'),
    'sell': ('السعر للعميل', 'sell', 'سعر العميل'),
    'invoice': ('بيان فاتورة', 'invoice'),
    'status': ('حالة التحصيل', 'status', 'الحالة'),
    'paid_date': ('تاريخ السداد', 'paid_date'),
    'pay_method': ('طريقة الدفع', 'payment_method'),
    'notes': ('ملحوظات', 'ملاحظات', 'notes'),
}


def _str(val) -> str:
    if val is None:
        return ''
    s = str(val).strip()
    return '' if s.lower() == 'nan' else s


def _float(val, default: float = 0.0) -> float:
    try:
        if val is None or _str(val) == '':
            return default
        return round(float(val), 2)
    except (TypeError, ValueError):
        return default


def _extract_cn(text: str) -> str | None:
    m = re.search(r'CN-\s*(\d+)', _str(text), re.I)
    if not m:
        return None
    return f'CN-{int(m.group(1)):05d}'


def _parse_date(val) -> date | None:
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = _str(val)
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _norm_op_number(val) -> str:
    s = _str(val)
    if not s:
        return ''
    digits = re.sub(r'\D', '', s)
    return digits.lstrip('0') or '0'


def _header_map(headers: list[str]) -> dict[str, int] | None:
    norm = {_str(h): i for i, h in enumerate(headers) if _str(h)}
    lower = {k.lower(): i for k, i in norm.items()}
    mapping: dict[str, int] = {}
    for key, names in COL_ALIASES.items():
        for name in names:
            idx = norm.get(name)
            if idx is None:
                idx = lower.get(name.lower())
            if idx is not None:
                mapping[key] = idx
                break
    if 'description' in mapping and 'date' in mapping:
        return mapping
    return None


def _cell(row: tuple, mapping: dict[str, int] | None, key: str, fallback: int) -> object:
    if mapping and key in mapping:
        idx = mapping[key]
        return row[idx] if idx < len(row) else None
    return row[fallback] if fallback < len(row) else None


def _row_record(row: tuple, col_map: dict[str, int] | None) -> dict:
    return {
        'title': _str(_cell(row, col_map, 'title', 0)),
        'op': _cell(row, col_map, 'op', 1),
        'contract': _str(_cell(row, col_map, 'contract', 2)),
        'date': _cell(row, col_map, 'date', 3),
        'description': _str(_cell(row, col_map, 'description', 4)),
        'cost': _cell(row, col_map, 'cost', 5),
        'sell': _cell(row, col_map, 'sell', 6),
        'invoice': _str(_cell(row, col_map, 'invoice', 7)),
        'status': _str(_cell(row, col_map, 'status', 8)),
        'paid_date': _cell(row, col_map, 'paid_date', 9),
        'pay_method': _str(_cell(row, col_map, 'pay_method', 10)),
        'notes': _str(_cell(row, col_map, 'notes', 11)),
    }


def _compose_notes(rec: dict) -> str:
    parts = []
    op = _norm_op_number(rec.get('op'))
    if op:
        parts.append(f'{OP_NOTE_PREFIX} {op.zfill(3)}')
    if rec.get('invoice'):
        parts.append(f"فاتورة: {rec['invoice']}")
    paid_on = _parse_date(rec.get('paid_date'))
    if paid_on:
        parts.append(f'تاريخ السداد: {paid_on.isoformat()}')
    if rec.get('notes'):
        parts.append(rec['notes'])
    return '\n'.join(parts)


def load_rows_from_workbook(wb) -> list[dict]:
    ws = wb.active
    raw = [tuple(r) for r in ws.iter_rows(values_only=True)]
    if not raw:
        return []
    header_idx = 0
    col_map: dict[str, int] | None = None
    for i, row in enumerate(raw):
        headers = [_str(v) for v in row]
        col_map = _header_map(headers)
        if col_map:
            header_idx = i
            break
        text = ' '.join(headers)
        if 'بيان قطع' in text or 'رقم العمل' in text or 'CN-' in text:
            header_idx = i
            break
    out = []
    for row in raw[header_idx + 1:]:
        if not any(_str(v) for v in row):
            continue
        rec = _row_record(row, col_map)
        if not _parse_date(rec.get('date')) or not rec.get('description'):
            continue
        out.append(rec)
    return out


def load_rows_from_bytes(data: bytes | BinaryIO) -> list[dict]:
    if isinstance(data, (bytes, bytearray)):
        stream = io.BytesIO(data)
    else:
        stream = data
    wb = openpyxl.load_workbook(stream, read_only=True, data_only=True)
    try:
        return load_rows_from_workbook(wb)
    finally:
        wb.close()


def load_rows(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return load_rows_from_workbook(wb)
    finally:
        wb.close()


def _existing_op_numbers() -> set[str]:
    found: set[str] = set()
    pattern = re.compile(r'رقم العملية:\s*(\d+)')
    for p in PartsBilling.query.all():
        if not p.notes:
            continue
        for m in pattern.finditer(p.notes):
            found.add(m.group(1).lstrip('0') or '0')
    return found


def import_parts_billing_rows(
    rows: list[dict],
    *,
    dry_run: bool = False,
    skip_existing: bool = True,
    uncollected_only: bool = False,
    force_uncollected: bool = False,
    db_session=None,
    next_code_fn=None,
) -> dict:
    stats = {
        'rows': len(rows),
        'imported': 0,
        'skipped_existing': 0,
        'skipped_missing': 0,
        'skipped_collected': 0,
        'errors': 0,
    }
    missing_samples: list[str] = []

    contracts = {
        c.code.upper(): c
        for c in Contract.query.all()
        if c.code
    }
    existing_ops = _existing_op_numbers() if skip_existing else set()

    for rec in rows:
        cn_code = _extract_cn(rec.get('contract')) or _extract_cn(rec.get('title'))
        billing_date = _parse_date(rec.get('date'))
        description = rec.get('description') or ''
        op_num = _norm_op_number(rec.get('op'))

        if not cn_code or not billing_date or not description:
            stats['errors'] += 1
            continue

        if skip_existing and op_num and op_num in existing_ops:
            stats['skipped_existing'] += 1
            continue

        contract = contracts.get(cn_code.upper())
        if not contract:
            stats['skipped_missing'] += 1
            if len(missing_samples) < 15:
                missing_samples.append(f'عملية {op_num or "?"}: عقد {cn_code} غير موجود')
            continue

        cost = _float(rec.get('cost'))
        sell = _float(rec.get('sell'))
        status = normalize_parts_status(rec.get('status') or '')
        if uncollected_only and status == 'محصل':
            stats['skipped_collected'] += 1
            continue
        if force_uncollected:
            status = 'غير محصل'
        payment_method = rec.get('pay_method') or ''
        notes = _compose_notes(rec)
        links = resolve_parts_links(contract_code=cn_code)

        if dry_run:
            stats['imported'] += 1
            if op_num:
                existing_ops.add(op_num)
            continue

        if not next_code_fn or not db_session:
            raise RuntimeError('next_code_fn and db_session required for import')

        part = PartsBilling(
            code=next_code_fn(PartsBilling, 'PB-', digits=3),
            customer_id=links['customer_id'],
            contract_id=links['contract_id'],
            elevator_id=links['elevator_id'],
            technician_id=links['technician_id'],
            visit_id=links['visit_id'],
            fault_id=links['fault_id'],
            billing_date=billing_date,
            description=description,
            cost_price=cost,
            sell_price=sell,
            profit=round(sell - cost, 2),
            paid_amount=sell if status == 'محصل' else 0,
            payment_method=payment_method,
            status=status,
            notes=notes or None,
        )
        db_session.add(part)
        if op_num:
            existing_ops.add(op_num)
        stats['imported'] += 1

    if not dry_run and db_session is not None:
        db_session.commit()

    stats['missing_samples'] = missing_samples
    return stats


def import_parts_billing_file(
    data: bytes | BinaryIO,
    *,
    dry_run: bool = False,
    skip_existing: bool = True,
    uncollected_only: bool = False,
    force_uncollected: bool = False,
    db_session=None,
    next_code_fn=None,
) -> dict:
    rows = load_rows_from_bytes(data)
    return import_parts_billing_rows(
        rows,
        dry_run=dry_run,
        skip_existing=skip_existing,
        uncollected_only=uncollected_only,
        force_uncollected=force_uncollected,
        db_session=db_session,
        next_code_fn=next_code_fn,
    )


def import_parts(
    path: str,
    *,
    dry_run: bool = False,
    skip_existing: bool = True,
    uncollected_only: bool = False,
    force_uncollected: bool = False,
) -> dict:
    """CLI helper — requires Flask app context with db imported."""
    from app import db, next_code

    rows = load_rows(path)
    return import_parts_billing_rows(
        rows,
        dry_run=dry_run,
        skip_existing=skip_existing,
        uncollected_only=uncollected_only,
        force_uncollected=force_uncollected,
        db_session=None if dry_run else db.session,
        next_code_fn=next_code,
    )
